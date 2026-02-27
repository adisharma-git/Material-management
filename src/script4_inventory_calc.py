"""
Script 4: Inventory Calculation Sheet Generator
Hospital Pharmacy Inventory Management
"""

import math
import os
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class InventoryCalculator:
    """Generates the complete Inventory Calculation sheet."""

    ITEM_COL = "Item\nCode"

    def __init__(self, log_fn=print):
        self.log = log_fn
        self.data: pd.DataFrame = None

    # ── Step 1 ────────────────────────────────────────────────────────────────────
    def load_master(self, filepath: str):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Master data not found: {filepath}")

        self.log("Loading Master Data…")
        df = pd.read_excel(filepath)
        total = len(df)
        df = df[df["Current SKU (TRUE/FALSE)"] == True].copy()
        self.log(f"  {len(df):,} active SKUs (of {total:,} total)")
        self.data = df

    # ── Step 2 ────────────────────────────────────────────────────────────────────
    def merge_global_stock(self, filepath: str):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Global stock lookup not found: {filepath}")

        self.log("Merging Global Stock…")
        gs = pd.read_excel(filepath, sheet_name="output")

        self.data = self.data.merge(
            gs,
            left_on=self.ITEM_COL,
            right_on="Item Code",
            how="left",
        )
        self.data["Global stock"] = self.data["Total Global Stock"].fillna(0)
        self.data.drop(columns=["Item Code", "Total Global Stock"], inplace=True)
        self.log(f"  Global stock merged – {(self.data['Global stock'] > 0).sum():,} items with stock")

    # ── Step 3 ────────────────────────────────────────────────────────────────────
    def merge_main_store_stock(self, filepath: str):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Main store lookup not found: {filepath}")

        self.log("Merging Main Store Stock…")
        ms = pd.read_excel(filepath, sheet_name="OUTPUT")

        self.data = self.data.merge(ms, on=self.ITEM_COL, how="left")
        self.data["Main Store Stock"] = self.data["Sum of Qty."].fillna(0)
        self.data.drop(columns=["Sum of Qty."], inplace=True)
        self.log(f"  Main store stock merged – {(self.data['Main Store Stock'] > 0).sum():,} items with stock")

    # ── Step 4 ────────────────────────────────────────────────────────────────────
    def process_pending_po(self, filepath: str):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Expected items file not found: {filepath}")

        self.log("Processing Pending PO…")

        # Detect CSV vs Excel
        if filepath.lower().endswith(".csv"):
            expected = pd.read_csv(filepath)
        else:
            expected = pd.read_excel(filepath)

        self.log(f"  Loaded {len(expected):,} PO lines")

        # Parse date
        expected["POCreated Date"] = pd.to_datetime(expected["POCreated Date"], errors="coerce")
        three_months_ago = datetime.now() - timedelta(days=90)

        # Filter
        expected = expected[expected["Store Name"] == "Main Medical Store (MMS)-SEPL"]
        expected = expected[expected["POCreated Date"] >= three_months_ago]
        self.log(f"  {len(expected):,} lines after store + 90-day filter")

        # Aggregate
        pending = expected.groupby("Item Code")["Pen.Qty"].sum().reset_index()
        pending.columns = ["Item Code", "Pending PO"]

        self.data = self.data.merge(
            pending,
            left_on=self.ITEM_COL,
            right_on="Item Code",
            how="left",
        )
        self.data["Pending PO"] = self.data["Pending PO"].fillna(0)
        self.data.drop(columns=["Item Code"], inplace=True)
        self.log(f"  Pending PO merged – {(self.data['Pending PO'] > 0).sum():,} items with pending orders")

    # ── Step 5-9 ──────────────────────────────────────────────────────────────────
    def calculate_metrics(self):
        self.log("Calculating inventory metrics…")
        df = self.data

        # Column R – Net Stock
        df["Net Stock"] = df["Global stock"] + df["Pending PO"]

        # Column S – Global Stock Days
        df["Global Stock Days"] = df.apply(
            lambda r: round(r["Global stock"] / r["ADC"], 0) if r["ADC"] > 0 else 0,
            axis=1,
        )

        # Column T – Main Store Stock Days
        df["Main Store Stock Days"] = df.apply(
            lambda r: round(r["Main Store Stock"] / r["ADC"], 0) if r["ADC"] > 0 else 0,
            axis=1,
        )

        # Column U – Reorder Needed?
        df["Reorder Needed?"] = (
            (df["Current SKU (TRUE/FALSE)"] == True)
            & (df["Net Stock"] < df["Min Stock Level"])
        )

        # Column V – Order Qty
        def calc_order_qty(row):
            if not row["Current SKU (TRUE/FALSE)"] or not row["Reorder Needed?"]:
                return 0
            shortage = row["Max Stock Level"] - row["Net Stock"]
            pack = row["Pack size"]
            if pack <= 0 or shortage <= 0:
                return 0
            return max(0, math.ceil(shortage / pack) * pack)

        df["Order Qty"] = df.apply(calc_order_qty, axis=1)
        self.data = df

    # ── Validate ──────────────────────────────────────────────────────────────────
    def validate(self):
        self.log("Validating…")
        df = self.data

        expected_cols = 22
        if len(df.columns) != expected_cols:
            self.log(f"  WARNING: Expected {expected_cols} columns, got {len(df.columns)}")
            self.log(f"  Columns: {df.columns.tolist()}")

        net_ok = (df["Net Stock"] == df["Global stock"] + df["Pending PO"]).all()
        assert net_ok, "Net Stock calculation error"
        assert (df["Order Qty"] >= 0).all(), "Negative order quantities found"
        self.log("  ✓ All validations passed")

    # ── Export ────────────────────────────────────────────────────────────────────
    def export(self, output_file: str):
        self.log(f"\nExporting to {output_file}…")

        # Sort: reorder items first, then by order qty desc
        df = self.data.sort_values(
            ["Reorder Needed?", "Order Qty"],
            ascending=[False, False],
        ).reset_index(drop=True)

        sheet_name = "Inventory Calculation"
        df.to_excel(output_file, sheet_name=sheet_name, index=False)

        # ── Formatting ────────────────────────────────────────────────────────────
        wb = load_workbook(output_file)
        ws = wb[sheet_name]

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Header style
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Highlight Reorder Needed column (U)
        reorder_col = None
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value == "Reorder Needed?":
                reorder_col = col[0].column
                break

        if reorder_col:
            red_fill   = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
            green_fill = PatternFill(start_color="D7FFD7", end_color="D7FFD7", fill_type="solid")
            for row in ws.iter_rows(2, ws.max_row, reorder_col, reorder_col):
                cell = row[0]
                cell.fill = red_fill if cell.value else green_fill

        # Column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_len = len(str(cell.value)) if cell.value is not None else 0
                    max_len = max(max_len, val_len)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

        wb.save(output_file)
        self.log(f"  ✓ Export complete with formatting")

    # ── Summary ───────────────────────────────────────────────────────────────────
    def print_summary(self):
        df = self.data
        unit_cost_col = next(
            (c for c in df.columns if "unit cost" in c.lower()), None
        )
        order_value = (
            (df["Order Qty"] * df[unit_cost_col]).sum()
            if unit_cost_col else 0
        )

        self.log("\n" + "=" * 55)
        self.log("INVENTORY CALCULATION – SUMMARY")
        self.log("=" * 55)
        self.log(f"  Total items processed  : {len(df):,}")
        self.log(f"  Items needing reorder  : {df['Reorder Needed?'].sum():,}")
        self.log(f"  Total order quantity   : {df['Order Qty'].sum():,.0f}")
        self.log(f"  Est. order value (₹)   : {order_value:,.2f}")
        self.log(f"  Avg global stock days  : {df['Global Stock Days'].mean():.1f}")
        self.log(f"  Avg main store days    : {df['Main Store Stock Days'].mean():.1f}")
        self.log("=" * 55)

    # ── Full run ──────────────────────────────────────────────────────────────────
    def run(
        self,
        master_file: str,
        global_stock_file: str,
        main_store_file: str,
        expected_items_file: str,
        output_file: str,
    ):
        self.load_master(master_file)
        self.merge_global_stock(global_stock_file)
        self.merge_main_store_stock(main_store_file)
        self.process_pending_po(expected_items_file)
        self.calculate_metrics()
        self.validate()
        self.export(output_file)
        self.print_summary()
        return self.data


if __name__ == "__main__":
    calc = InventoryCalculator()
    calc.run(
        master_file="MASTER_DATA_INPUT.xlsx",
        global_stock_file="Material_Global_Stock_Lookup.xlsx",
        main_store_file="Material_Main_Store_Stock_Lookup.xlsx",
        expected_items_file="Expected_Items_Material.xlsx",
        output_file="INVENTORY_CALCULATION.xlsx",
    )
